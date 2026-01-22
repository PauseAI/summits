#!/usr/bin/env python3
"""
Generate contacts.xlsx from contacts.json

Usage: python generate_xlsx.py [--output path/to/output.xlsx]
"""

import json
import argparse
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed. Run: pip install openpyxl")
    exit(1)


def load_data():
    """Load contacts.json and summits.json"""
    data_dir = Path(__file__).parent.parent / "data"
    
    with open(data_dir / "contacts.json", "r") as f:
        contacts = json.load(f)
    
    with open(data_dir / "summits.json", "r") as f:
        summits = json.load(f)
    
    return contacts, summits


def create_nations_sheet(ws, contacts, summits):
    """Create the Nations - AI Summits sheet"""
    headers = [
        "Nation", "Region", "GPAI Member", "OECD Member",
        "Bletchley 2023", "Seoul 2024", "Paris 2025",
        "AI Safety Institute", "Notes"
    ]
    
    # Styles
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data
    row = 2
    for country_key, country in contacts.get("countries", {}).items():
        summits_data = country.get("summits", {})
        
        ws.cell(row=row, column=1, value=country.get("name", country_key))
        ws.cell(row=row, column=2, value=country.get("region", ""))
        ws.cell(row=row, column=3, value="Y" if country.get("gpai_member") else "N")
        ws.cell(row=row, column=4, value="Y" if country.get("oecd_member") else "N")
        ws.cell(row=row, column=5, value=summits_data.get("bletchley_2023", {}).get("status", ""))
        ws.cell(row=row, column=6, value=summits_data.get("seoul_2024", {}).get("status", ""))
        ws.cell(row=row, column=7, value=summits_data.get("paris_2025", {}).get("status", ""))
        
        aisi = country.get("ai_safety_institute", {})
        if isinstance(aisi, dict):
            ws.cell(row=row, column=8, value="Y" if aisi.get("exists") else "N")
        else:
            ws.cell(row=row, column=8, value="Y" if aisi else "N")
        
        ws.cell(row=row, column=9, value=country.get("notes", ""))
        
        # Apply border to all cells
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    # Adjust column widths
    widths = [20, 12, 12, 12, 14, 12, 12, 16, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"


def create_contacts_sheet(ws, contacts):
    """Create the Contacts sheet"""
    headers = [
        "Nation", "Name", "Title", "Organization", "Role Type",
        "Email", "Source URL", "Verified Date", "Notes"
    ]
    
    # Styles
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = thin_border
    
    # Write data
    row = 2
    for country_key, country in contacts.get("countries", {}).items():
        country_name = country.get("name", country_key)
        
        for contact in country.get("contacts", []):
            ws.cell(row=row, column=1, value=country_name)
            ws.cell(row=row, column=2, value=contact.get("name", ""))
            ws.cell(row=row, column=3, value=contact.get("title", ""))
            ws.cell(row=row, column=4, value=contact.get("organization", ""))
            ws.cell(row=row, column=5, value=contact.get("role_type", ""))
            ws.cell(row=row, column=6, value=contact.get("email", ""))
            ws.cell(row=row, column=7, value=contact.get("source_url", ""))
            ws.cell(row=row, column=8, value=contact.get("verified_date", ""))
            ws.cell(row=row, column=9, value=contact.get("notes", ""))
            
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Adjust column widths
    widths = [18, 20, 35, 20, 12, 30, 40, 12, 40]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.freeze_panes = "A2"


def create_summits_sheet(ws, summits):
    """Create the Summits sheet"""
    headers = ["Summit", "Location", "Dates", "Host", "Declaration", "Notes"]
    
    # Styles
    header_fill = PatternFill("solid", fgColor="6A1B9A")
    header_font = Font(bold=True, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    row = 2
    for summit_key, summit in summits.get("summits", {}).items():
        dates = summit.get("dates", {})
        date_str = f"{dates.get('start', '')} to {dates.get('end', '')}" if dates else ""
        
        ws.cell(row=row, column=1, value=summit.get("name", summit_key))
        ws.cell(row=row, column=2, value=summit.get("location", ""))
        ws.cell(row=row, column=3, value=date_str)
        ws.cell(row=row, column=4, value=summit.get("host_country", ""))
        ws.cell(row=row, column=5, value=summit.get("declaration", {}).get("name", "") if isinstance(summit.get("declaration"), dict) else "")
        ws.cell(row=row, column=6, value=summit.get("notes", ""))
        
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = thin_border
        
        row += 1
    
    widths = [25, 30, 25, 18, 30, 50]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    ws.freeze_panes = "A2"


def main():
    parser = argparse.ArgumentParser(description="Generate contacts spreadsheet from JSON")
    parser.add_argument("--output", "-o", default="data/contacts.xlsx", help="Output path")
    args = parser.parse_args()
    
    print("Loading data...")
    contacts, summits = load_data()
    
    print("Creating workbook...")
    wb = Workbook()
    
    # Remove default sheet and create our sheets
    wb.remove(wb.active)
    
    ws_nations = wb.create_sheet("Nations - AI Summits")
    create_nations_sheet(ws_nations, contacts, summits)
    
    ws_contacts = wb.create_sheet("Contacts")
    create_contacts_sheet(ws_contacts, contacts)
    
    ws_summits = wb.create_sheet("Summits")
    create_summits_sheet(ws_summits, summits)
    
    # Save
    output_path = Path(__file__).parent.parent / args.output
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"Saved to {output_path}")


if __name__ == "__main__":
    main()
